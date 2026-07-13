#!/usr/bin/env python3
"""
scripts/sync_from_sharepoint.py

Pulls Client ID list.xlsx from the local SharePoint OneDrive sync folder,
sanitises it (keeps only the 'Client Info' sheet) into data/instances.xlsx,
regenerates the JSON files, and pushes the result to GitHub.

Designed to be run unattended by a Windows Scheduled Task.

Usage:
    python scripts/sync_from_sharepoint.py [--dry-run]

Flags:
    --dry-run    Sanitise + regenerate JSON locally but do NOT git add/commit/push.
                 Useful for validating column/sheet structure after upstream edits.
"""

import argparse
import logging
import logging.handlers
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

SOURCE_PATH = Path(
    r"C:\Users\xavier.parro\GXO\GXO Blue Yonder - Documentos\BY Client Tracker\Client ID list.xlsx"
)
REPO_ROOT = Path(
    r"C:\Users\xavier.parro\OneDrive - GXO\Documentos\GitHub\GXO\instances-manager"
)
TARGET_XLSX = REPO_ROOT / "data" / "instances.xlsx"
LOG_FILE = REPO_ROOT / "scripts" / "sync_from_sharepoint.log"
CLIENT_INFO_SHEET = "Client Info"

# A Scheduled Task runs with a minimal environment where git may not be on PATH,
# so resolve it to an absolute path and fail loudly if it cannot be found.
GIT = shutil.which("git") or r"C:\Program Files\Git\cmd\git.exe"

# Only these column headers are operational and safe to publish. Anything else in
# the source (T&E, invoices, GXO approvals, internal notes, salesforce cases, etc.)
# is stripped before the sanitised copy is committed to the repo.
KEEP_HEADERS = {
    "POC",
    "Client ID",
    "Client",
    "NP Env",
    "Prod Env",
    "Arch Env",
    "Site Code",
    "Instance",
    "NP BY Version",
    "Prod. BY Version",
    "NP Realm",
    "GIT Repo",
    "Prod. Realm",
    "Location",
    "Operating Hours",
    "SIte distribution",
    "Site Contact",
    "Go-Live",
}

_FORMAT = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
_rotating = logging.handlers.RotatingFileHandler(
    str(LOG_FILE), maxBytes=1_000_000, backupCount=3, encoding="utf-8"
)
_rotating.setFormatter(_FORMAT)
_console = logging.StreamHandler(sys.stdout)
_console.setFormatter(_FORMAT)
logging.basicConfig(level=logging.INFO, handlers=[_rotating, _console])


def git(*args):
    return run([GIT, *args])


def run(cmd):
    logging.info("$ %s", " ".join(str(c) for c in cmd))
    result = subprocess.run(
        cmd, capture_output=True, text=True, cwd=str(REPO_ROOT), encoding="utf-8"
    )
    if result.stdout and result.stdout.strip():
        logging.info("stdout: %s", result.stdout.strip())
    if result.stderr and result.stderr.strip():
        logging.info("stderr: %s", result.stderr.strip())
    return result


def sanitise_xlsx():
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Source not found: {SOURCE_PATH}")

    logging.info("Reading source: %s", SOURCE_PATH)
    wb = openpyxl.load_workbook(str(SOURCE_PATH), data_only=True)

    if CLIENT_INFO_SHEET not in wb.sheetnames:
        raise ValueError(
            f"'{CLIENT_INFO_SHEET}' sheet missing. Sheets present: {wb.sheetnames}"
        )

    for sheet_name in list(wb.sheetnames):
        if sheet_name != CLIENT_INFO_SHEET:
            del wb[sheet_name]

    ws = wb[CLIENT_INFO_SHEET]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    drop_indices = [
        idx for idx, h in enumerate(headers, start=1) if h not in KEEP_HEADERS
    ]
    missing = KEEP_HEADERS - {h for h in headers if h in KEEP_HEADERS}
    if missing:
        raise ValueError(
            f"Expected headers missing from source sheet: {sorted(missing)}. "
            f"Headers present: {headers}"
        )
    for idx in sorted(drop_indices, reverse=True):
        ws.delete_cols(idx)
    logging.info(
        "Kept %d columns; dropped %d sensitive/unused columns",
        len(headers) - len(drop_indices),
        len(drop_indices),
    )

    TARGET_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(TARGET_XLSX))
    logging.info("Wrote sanitised copy: %s", TARGET_XLSX)


def regenerate_json():
    result = run([sys.executable, "scripts/excel_to_json.py", "data/instances.xlsx"])
    if result.returncode != 0:
        raise RuntimeError(f"excel_to_json.py exited {result.returncode}")


def git_data_dirty():
    result = git("status", "--porcelain", "data/")
    return bool(result.stdout.strip())


_TIMESTAMP_FIELD = re.compile(r'"(generated_at|last_updated)":')


def only_timestamp_changes_in_json():
    """True iff data/instances.json differs from HEAD only on timestamp lines.

    The XLSX always re-serialises with a fresh modification timestamp and the JSON
    re-stamps generated_at/last_updated on every run, so a byte-level dirty check
    over-reports changes. We use the JSON diff (text, easy to inspect) as the
    source of truth for 'did real data change?'.
    """
    result = git("diff", "--unified=0", "--", "data/instances.json")
    if not result.stdout.strip():
        return True
    for line in result.stdout.splitlines():
        if line.startswith(("+++", "---", "@@", "diff ", "index ")):
            continue
        if line.startswith(("+", "-")):
            if not _TIMESTAMP_FIELD.search(line):
                return False
    return True


def commit_and_push():
    git(
        "add",
        "data/instances.xlsx",
        "data/instances.json",
        "data/instances.min.json",
    )
    msg = f"Auto-sync from SharePoint - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    commit = git("commit", "-m", msg)
    if commit.returncode != 0:
        if "nothing to commit" in (commit.stdout or "").lower():
            logging.info("Nothing to commit after staging")
            return
        raise RuntimeError(f"git commit failed (exit {commit.returncode})")

    # Edits made directly on github.com (index.html, styling, ...) land on origin/main
    # without this clone ever seeing them, which makes the push a non-fast-forward.
    # Replay our data commit on top of whatever is upstream before pushing.
    rebase = git("pull", "--rebase", "origin", "main")
    if rebase.returncode != 0:
        git("rebase", "--abort")
        raise RuntimeError(
            f"git pull --rebase failed (exit {rebase.returncode}); "
            "resolve the divergence manually"
        )

    push = git("push", "origin", "main")
    if push.returncode != 0:
        raise RuntimeError(f"git push failed (exit {push.returncode})")
    logging.info("Pushed to origin/main")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Sanitise + regenerate JSON but do not commit/push.",
    )
    args = parser.parse_args()

    logging.info("=" * 60)
    logging.info("Sync run started%s", " (DRY-RUN)" if args.dry_run else "")

    try:
        sanitise_xlsx()
        regenerate_json()

        if not git_data_dirty():
            logging.info("No changes in data/ after regen. Nothing to do.")
            return 0

        if only_timestamp_changes_in_json():
            logging.info(
                "Only timestamp fields changed. Reverting data/ and skipping commit."
            )
            git("checkout", "--", "data/")
            return 0

        if args.dry_run:
            logging.info("DRY-RUN: real data changes detected; skipping commit/push.")
            return 0

        commit_and_push()
        logging.info("Sync completed OK")
        return 0
    except Exception:
        logging.exception("Sync FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(main())
